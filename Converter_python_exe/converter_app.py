import sys
import os
import pandas as pd
from datetime import datetime
from PySide6.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
    QLabel, QLineEdit, QPushButton, QFileDialog, QMessageBox,
    QGroupBox, QRadioButton, QButtonGroup, QProgressBar, QTextEdit
)
from PySide6.QtCore import Qt, QThread, Signal
from PySide6.QtGui import QFont, QPalette, QColor

class ConverterThread(QThread):
    """Поток для выполнения конвертации"""
    progress = Signal(int)
    message = Signal(str)
    finished = Signal(bool, str)
    
    def __init__(self, input_file, output_file, is_xlsx):
        super().__init__()
        self.input_file = input_file
        self.output_file = output_file
        self.is_xlsx = is_xlsx
    
    def run(self):
        try:
            self.message.emit("Начало обработки файла...")
            
            # Читаем и обрабатываем данные
            data = []
            total_lines = 0
            
            # Считаем количество строк для прогресса
            with open(self.input_file, 'r', encoding='utf-8') as f:
                for _ in f:
                    total_lines += 1
            
            if total_lines == 0:
                self.finished.emit(False, "Файл пуст")
                return
            
            self.message.emit(f"Найдено строк: {total_lines}")
            
            processed = 0
            with open(self.input_file, 'r', encoding='utf-8') as f:
                for line in f:
                    line = line.strip()
                    if line and not line.startswith('#'):
                        parts = line.split(';')
                        if len(parts) >= 6:
                            # Обработка чисел
                            for i in range(len(parts)):
                                if i == 0:  # Time_ms
                                    parts[i] = self.remove_leading_zeros(parts[i])
                                elif 1 <= i <= 3:  # PITCH, ROLL, YAW
                                    parts[i] = parts[i].replace('.', ',')
                                    parts[i] = self.remove_leading_zeros_decimal(parts[i])
                            
                            data.append(parts[:6])
                    
                    processed += 1
                    progress = int((processed / total_lines) * 100)
                    self.progress.emit(progress)
            
            if not data:
                self.finished.emit(False, "Нет данных для обработки")
                return
            
            self.message.emit("Создание DataFrame...")
            df = pd.DataFrame(data, columns=['Time_ms', 'PITCH', 'ROLL', 'YAW', 'Dizziness', 'Nystagmus'])
            
            if self.is_xlsx:
                self.message.emit("Сохранение в XLSX...")
                with pd.ExcelWriter(self.output_file, engine='openpyxl') as writer:
                    df.to_excel(writer, index=False, sheet_name='Data')
                    
                    workbook = writer.book
                    worksheet = writer.sheets['Data']
                    
                    from openpyxl.styles import Font, Alignment, PatternFill
                    from openpyxl.utils import get_column_letter
                    
                    header_fill = PatternFill(start_color='DCE6F1', end_color='DCE6F1', fill_type='solid')
                    header_font = Font(bold=True, size=12)
                    header_alignment = Alignment(horizontal='center', vertical='center')
                    
                    for col in range(1, len(df.columns) + 1):
                        cell = worksheet.cell(row=1, column=col)
                        cell.fill = header_fill
                        cell.font = header_font
                        cell.alignment = header_alignment
                        
                        column_letter = get_column_letter(col)
                        worksheet.column_dimensions[column_letter].auto_size = True
                
                file_size = os.path.getsize(self.output_file)
                message = f"✅ ФАЙЛ УСПЕШНО ПРЕОБРАЗОВАН!\n\n" \
                         f"Сохранен как: {os.path.basename(self.output_file)}\n" \
                         f"Размер: {file_size:,} байт\n" \
                         f"Строк данных: {len(data)}\n\n" \
                         f"Файл готов к открытию в Microsoft Excel."
            
            else:
                self.message.emit("Сохранение в CSV...")
                df.to_csv(self.output_file, sep=';', index=False, encoding='utf-8-sig')
                
                file_size = os.path.getsize(self.output_file)
                message = f"✅ ФАЙЛ УСПЕШНО ПРЕОБРАЗОВАН!\n\n" \
                         f"Сохранен как: {os.path.basename(self.output_file)}\n" \
                         f"Размер: {file_size:,} байт\n" \
                         f"Строк данных: {len(data)}\n\n" \
                         f"При открытии в Excel:\n" \
                         f"1. Выберите 'Все файлы (*.*)'\n" \
                         f"2. Укажите кодировку UTF-8\n" \
                         f"3. Выберите разделитель ';'"
            
            self.finished.emit(True, message)
            
        except Exception as e:
            error_msg = f"Ошибка при конвертации:\n{str(e)}"
            self.finished.emit(False, error_msg)
    
    def remove_leading_zeros(self, s):
        if not s:
            return s
        while len(s) > 1 and s.startswith('0') and not s.startswith('0.'):
            s = s[1:]
        return s
    
    def remove_leading_zeros_decimal(self, s):
        if not s:
            return s
        
        is_negative = False
        if s.startswith('-'):
            is_negative = True
            s = s[1:]
        
        if '.' in s or ',' in s:
            separator = '.' if '.' in s else ','
            parts = s.split(separator)
            
            if len(parts) == 2:
                integer_part = parts[0]
                decimal_part = parts[1]
                
                while len(integer_part) > 1 and integer_part.startswith('0'):
                    integer_part = integer_part[1:]
                
                s = integer_part + separator + decimal_part
        
        if is_negative:
            s = '-' + s
        
        return s

class ConverterApp(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Конвертер txt в Excel v1.0")
        self.setGeometry(100, 100, 800, 500)
        
        self.setup_dark_theme()
        
        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        
        main_layout = QVBoxLayout(central_widget)
        main_layout.setSpacing(15)
        main_layout.setContentsMargins(20, 20, 20, 20)
        
        # === ЗАГОЛОВОК ===
        title_label = QLabel("Конвертер txt в Excel")
        title_label.setAlignment(Qt.AlignCenter)
        title_label.setFont(QFont("Arial", 20, QFont.Bold))
        title_label.setStyleSheet("color: #4CAF50;")
        main_layout.addWidget(title_label)
        
        subtitle_label = QLabel("Преобразователь текстовых файлов программы MonitorHead в формат Excel")
        subtitle_label.setAlignment(Qt.AlignCenter)
        subtitle_label.setStyleSheet("color: #aaa; font-size: 12px;")
        main_layout.addWidget(subtitle_label)
        
        # === БЛОК ВЫБОРА ФАЙЛА ===
        file_group = QGroupBox("Исходный файл (.txt)")
        file_group.setStyleSheet("""
            QGroupBox {
                font-weight: bold;
                color: #4CAF50;
                border: 1px solid #444;
                border-radius: 5px;
                margin-top: 10px;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                left: 10px;
                padding: 0 5px 0 5px;
            }
        """)
        
        file_layout = QVBoxLayout()
        file_row = QHBoxLayout()
        file_row.addWidget(QLabel("Файл:"))
        
        self.input_file_edit = QLineEdit()
        self.input_file_edit.setPlaceholderText("Выберите файл .txt")
        self.input_file_edit.setReadOnly(True)
        self.input_file_edit.setStyleSheet("""
            QLineEdit {
                background-color: #3c3c3c;
                border: 1px solid #555;
                border-radius: 3px;
                padding: 5px;
                color: white;
            }
        """)
        file_row.addWidget(self.input_file_edit, 1)
        
        self.browse_file_btn = QPushButton("📁 Обзор")
        self.browse_file_btn.setFixedWidth(100)
        self.browse_file_btn.clicked.connect(self.browse_input_file)
        self.browse_file_btn.setStyleSheet(self.get_button_style())
        file_row.addWidget(self.browse_file_btn)
        
        file_layout.addLayout(file_row)
        file_group.setLayout(file_layout)
        main_layout.addWidget(file_group)
        
        # === ПЕРЕКЛЮЧАТЕЛЬ ФОРМАТА ===
        format_group = QGroupBox("Формат выходного файла")
        format_group.setStyleSheet("""
            QGroupBox {
                font-weight: bold;
                color: #4CAF50;
                border: 1px solid #444;
                border-radius: 5px;
                margin-top: 10px;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                left: 10px;
                padding: 0 5px 0 5px;
            }
        """)
        
        format_layout = QHBoxLayout()
        
        # Создаем радиокнопки (XLSX выбран по умолчанию)
        self.xlsx_radio = QRadioButton("XLSX (нативный формат Excel)")
        self.csv_radio = QRadioButton("CSV (универсальный текстовый формат)")
        self.xlsx_radio.setChecked(True)  # XLSX выбран по умолчанию
        
        # Устанавливаем начальные стили для радиокнопок
        self.update_radio_styles()
        
        self.format_group = QButtonGroup()
        self.format_group.addButton(self.xlsx_radio)
        self.format_group.addButton(self.csv_radio)
        
        # Подключаем сигналы изменения состояния
        self.xlsx_radio.toggled.connect(self.on_format_changed)
        self.csv_radio.toggled.connect(self.on_format_changed)
        
        format_layout.addWidget(self.xlsx_radio)
        format_layout.addWidget(self.csv_radio)
        format_layout.addStretch()
        
        format_group.setLayout(format_layout)
        main_layout.addWidget(format_group)
        
        # === БЛОК НАСТРОЕК ВЫХОДНОГО ФАЙЛА ===
        output_group = QGroupBox("Выходной файл")
        output_group.setStyleSheet("""
            QGroupBox {
                font-weight: bold;
                color: #4CAF50;
                border: 1px solid #444;
                border-radius: 5px;
                margin-top: 10px;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                left: 10px;
                padding: 0 5px 0 5px;
            }
        """)
        
        output_layout = QVBoxLayout()
        
        # Путь сохранения
        path_row = QHBoxLayout()
        path_row.addWidget(QLabel("Папка сохранения:"))
        
        self.output_dir_edit = QLineEdit()
        self.output_dir_edit.setPlaceholderText("Папка для сохранения")
        self.output_dir_edit.setStyleSheet("""
            QLineEdit {
                background-color: #3c3c3c;
                border: 1px solid #555;
                border-radius: 3px;
                padding: 5px;
                color: white;
            }
        """)
        path_row.addWidget(self.output_dir_edit, 1)
        
        self.browse_dir_btn = QPushButton("📁 Обзор")
        self.browse_dir_btn.setFixedWidth(100)
        self.browse_dir_btn.clicked.connect(self.browse_output_dir)
        self.browse_dir_btn.setStyleSheet(self.get_button_style())
        path_row.addWidget(self.browse_dir_btn)
        
        output_layout.addLayout(path_row)
        
        # Имя файла
        name_row = QHBoxLayout()
        name_row.addWidget(QLabel("Имя файла:"))
        
        self.output_name_edit = QLineEdit()
        self.output_name_edit.setPlaceholderText("Имя будет взято из входного файла")
        self.output_name_edit.setStyleSheet("""
            QLineEdit {
                background-color: #3c3c3c;
                border: 1px solid #555;
                border-radius: 3px;
                padding: 5px;
                color: white;
            }
        """)
        name_row.addWidget(self.output_name_edit, 1)
        
        output_layout.addLayout(name_row)
        output_group.setLayout(output_layout)
        main_layout.addWidget(output_group)
        
        # === КНОПКА ПРЕОБРАЗОВАНИЯ (фиксированной ширины) ===
        self.convert_btn = QPushButton("ПРЕОБРАЗОВАТЬ В XLSX")
        self.convert_btn.setFixedSize(300, 45)  # Фиксированная ширина и высота
        self.convert_btn.setFont(QFont("Arial", 12, QFont.Bold))
        self.convert_btn.clicked.connect(self.start_conversion)
        self.convert_btn.setEnabled(False)
        self.convert_btn.setStyleSheet(self.get_convert_button_style(True))  # XLSX стиль по умолчанию
        main_layout.addWidget(self.convert_btn, alignment=Qt.AlignCenter)
        
        # === ПРОГРЕСС БАР ===
        self.progress_bar = QProgressBar()
        self.progress_bar.setVisible(False)
        self.progress_bar.setStyleSheet("""
            QProgressBar {
                border: 1px solid #444;
                border-radius: 3px;
                text-align: center;
                color: white;
            }
            QProgressBar::chunk {
                background-color: #4CAF50;
                border-radius: 3px;
            }
        """)
        main_layout.addWidget(self.progress_bar)
        
        # === СТАТУС ===
        self.status_label = QLabel("Готов к работе")
        self.status_label.setAlignment(Qt.AlignCenter)
        self.status_label.setStyleSheet("color: #aaa; font-size: 14px;")
        main_layout.addWidget(self.status_label)
        
        # === ИНФОРМАЦИЯ О ПРОГРАММЕ ===
        info_label = QLabel("© 2026 Конвертер txt в Excel v1.0")
        info_label.setAlignment(Qt.AlignCenter)
        info_label.setStyleSheet("color: #666; font-size: 10px;")
        main_layout.addWidget(info_label)
        
        # Инициализируем переменную для хранения базового имени файла
        self.base_file_name = ""
        
        # Подключаем сигналы изменения формата
        self.xlsx_radio.toggled.connect(self.update_file_extension)
        self.csv_radio.toggled.connect(self.update_file_extension)
        
        self.converter_thread = None
    
    def setup_dark_theme(self):
        """Настройка темной темы"""
        palette = QPalette()
        palette.setColor(QPalette.Window, QColor(30, 30, 30))
        palette.setColor(QPalette.WindowText, Qt.white)
        palette.setColor(QPalette.Base, QColor(45, 45, 45))
        palette.setColor(QPalette.AlternateBase, QColor(53, 53, 53))
        palette.setColor(QPalette.ToolTipBase, Qt.white)
        palette.setColor(QPalette.ToolTipText, Qt.white)
        palette.setColor(QPalette.Text, Qt.white)
        palette.setColor(QPalette.Button, QColor(53, 53, 53))
        palette.setColor(QPalette.ButtonText, Qt.white)
        palette.setColor(QPalette.BrightText, Qt.red)
        palette.setColor(QPalette.Link, QColor(42, 130, 218))
        palette.setColor(QPalette.Highlight, QColor(42, 130, 218))
        palette.setColor(QPalette.HighlightedText, Qt.black)
        
        self.setPalette(palette)
    
    def get_button_style(self, hover_color="#505050", pressed_color="#303030"):
        """Стиль для обычных кнопок"""
        return f"""
            QPushButton {{
                background-color: #404040;
                color: white;
                border: none;
                border-radius: 4px;
                padding: 8px 15px;
                font-weight: bold;
            }}
            QPushButton:hover {{
                background-color: {hover_color};
            }}
            QPushButton:pressed {{
                background-color: {pressed_color};
            }}
            QPushButton:disabled {{
                background-color: #2a2a2a;
                color: #888;
            }}
        """
    
    def get_convert_button_style(self, is_xlsx):
        """Стиль для кнопки преобразования (одинаковая ширина для обоих вариантов)"""
        if is_xlsx:
            return """
                QPushButton {
                    background-color: #3498db;
                    color: white;
                    border: none;
                    border-radius: 6px;
                    padding: 12px 20px;
                    font-weight: bold;
                    min-width: 300px;
                }
                QPushButton:hover {
                    background-color: #2e86c1;
                }
                QPushButton:pressed {
                    background-color: #1a5276;
                }
                QPushButton:disabled {
                    background-color: #555;
                    color: #888;
                }
            """
        else:
            return """
                QPushButton {
                    background-color: #4CAF50;
                    color: white;
                    border: none;
                    border-radius: 6px;
                    padding: 12px 20px;
                    font-weight: bold;
                    min-width: 300px;
                }
                QPushButton:hover {
                    background-color: #45a049;
                }
                QPushButton:pressed {
                    background-color: #3a5c42;
                }
                QPushButton:disabled {
                    background-color: #555;
                    color: #888;
                }
            """
    
    def update_radio_styles(self):
        """Обновляет стили радиокнопок в зависимости от выбранной"""
        if self.xlsx_radio.isChecked():
            # XLSX выбран - синий
            self.xlsx_radio.setStyleSheet("""
                QRadioButton {
                    color: #3498db;
                    font-weight: bold;
                    font-size: 13px;
                }
                QRadioButton::indicator {
                    width: 16px;
                    height: 16px;
                }
                QRadioButton::indicator:checked {
                    background-color: #3498db;
                    border: 3px solid #3498db;
                    border-radius: 8px;
                }
                QRadioButton::indicator:unchecked {
                    border: 2px solid #666;
                    border-radius: 8px;
                    background-color: #2a2a2a;
                }
            """)
            
            # CSV не выбран - серый
            self.csv_radio.setStyleSheet("""
                QRadioButton {
                    color: #888;
                    font-size: 13px;
                }
                QRadioButton::indicator {
                    width: 16px;
                    height: 16px;
                }
                QRadioButton::indicator:checked {
                    background-color: #666;
                    border: 3px solid #666;
                    border-radius: 8px;
                }
                QRadioButton::indicator:unchecked {
                    border: 2px solid #555;
                    border-radius: 8px;
                    background-color: #2a2a2a;
                }
            """)
        else:
            # CSV выбран - зеленый
            self.csv_radio.setStyleSheet("""
                QRadioButton {
                    color: #4CAF50;
                    font-weight: bold;
                    font-size: 13px;
                }
                QRadioButton::indicator {
                    width: 16px;
                    height: 16px;
                }
                QRadioButton::indicator:checked {
                    background-color: #4CAF50;
                    border: 3px solid #4CAF50;
                    border-radius: 8px;
                }
                QRadioButton::indicator:unchecked {
                    border: 2px solid #666;
                    border-radius: 8px;
                    background-color: #2a2a2a;
                }
            """)
            
            # XLSX не выбран - серый
            self.xlsx_radio.setStyleSheet("""
                QRadioButton {
                    color: #888;
                    font-size: 13px;
                }
                QRadioButton::indicator {
                    width: 16px;
                    height: 16px;
                }
                QRadioButton::indicator:checked {
                    background-color: #666;
                    border: 3px solid #666;
                    border-radius: 8px;
                }
                QRadioButton::indicator:unchecked {
                    border: 2px solid #555;
                    border-radius: 8px;
                    background-color: #2a2a2a;
                }
            """)
    
    def on_format_changed(self):
        """Вызывается при изменении формата"""
        # Обновляем стили радиокнопок
        self.update_radio_styles()
        
        # Обновляем стиль кнопки преобразования
        is_xlsx = self.xlsx_radio.isChecked()
        self.convert_btn.setStyleSheet(self.get_convert_button_style(is_xlsx))
        
        # Обновляем текст кнопки (без эмодзи ракеты)
        if is_xlsx:
            self.convert_btn.setText("ПРЕОБРАЗОВАТЬ В XLSX")
        else:
            self.convert_btn.setText("ПРЕОБРАЗОВАТЬ В CSV")
        
        # Обновляем расширение файла если имя уже задано
        self.update_file_extension()
    
    def browse_input_file(self):
        """Выбор входного файла"""
        file_name, _ = QFileDialog.getOpenFileName(
            self, "Выберите текстовый файл", 
            "", "Текстовые файлы (*.txt);;Все файлы (*.*)"
        )
        
        if file_name:
            self.input_file_edit.setText(file_name)
            
            # Сохраняем путь к папке и базовое имя файла
            dir_path = os.path.dirname(file_name)
            full_name = os.path.basename(file_name)
            
            # Сохраняем базовое имя без расширения
            if full_name.lower().endswith('.txt'):
                self.base_file_name = full_name[:-4]  # Убираем .txt
            else:
                self.base_file_name = full_name
            
            # Устанавливаем папку сохранения (та же что и у исходного файла)
            self.output_dir_edit.setText(dir_path)
            
            # Устанавливаем имя выходного файла с правильным расширением
            self.update_file_extension()
            
            # Проверяем, можно ли активировать кнопку преобразования
            self.check_convert_button()
            
            # Обновляем статус
            self.status_label.setText(f"Выбран файл: {full_name}")
            self.status_label.setStyleSheet("color: #aaa; font-size: 14px;")
    
    def browse_output_dir(self):
        """Выбор папки для сохранения"""
        dir_path = QFileDialog.getExistingDirectory(
            self, "Выберите папку для сохранения", ""
        )
        
        if dir_path:
            self.output_dir_edit.setText(dir_path)
            self.check_convert_button()
    
    def update_file_extension(self):
        """Обновление расширения файла при изменении формата"""
        if not self.base_file_name:
            return
        
        # Определяем расширение в зависимости от выбранного формата
        if self.xlsx_radio.isChecked():
            extension = ".xlsx"
        else:
            extension = ".csv"
        
        # Формируем полное имя файла
        new_name = self.base_file_name + extension
        
        # Устанавливаем в поле
        self.output_name_edit.setText(new_name)
        
        # Обновляем цвет текста в зависимости от формата
        if self.xlsx_radio.isChecked():
            self.output_name_edit.setStyleSheet("""
                QLineEdit {
                    background-color: #3c3c3c;
                    border: 1px solid #3498db;
                    border-radius: 3px;
                    padding: 5px;
                    color: #3498db;
                    font-weight: bold;
                }
            """)
        else:
            self.output_name_edit.setStyleSheet("""
                QLineEdit {
                    background-color: #3c3c3c;
                    border: 1px solid #4CAF50;
                    border-radius: 3px;
                    padding: 5px;
                    color: #4CAF50;
                    font-weight: bold;
                }
            """)
    
    def check_convert_button(self):
        """Активация кнопки преобразования при заполнении всех полей"""
        has_input = bool(self.input_file_edit.text())
        has_output_dir = bool(self.output_dir_edit.text())
        has_output_name = bool(self.output_name_edit.text())
        
        self.convert_btn.setEnabled(has_input and has_output_dir and has_output_name)
    
    def start_conversion(self):
        """Запуск процесса конвертации"""
        input_file = self.input_file_edit.text()
        output_dir = self.output_dir_edit.text()
        output_name = self.output_name_edit.text()
        is_xlsx = self.xlsx_radio.isChecked()
        
        # Формируем полный путь к выходному файлу
        output_file = os.path.join(output_dir, output_name)
        
        # Проверяем существование входного файла
        if not os.path.exists(input_file):
            QMessageBox.critical(self, "Ошибка", "Входной файл не существует!")
            return
        
        # Проверяем существование папки для сохранения
        if not os.path.exists(output_dir):
            try:
                os.makedirs(output_dir)
            except Exception as e:
                QMessageBox.critical(self, "Ошибка", f"Не удалось создать папку:\n{str(e)}")
                return
        
        # Проверяем, не совпадают ли имена входного и выходного файлов
        if os.path.normpath(input_file) == os.path.normpath(output_file):
            reply = QMessageBox.question(
                self, "Подтверждение",
                "Входной и выходной файлы совпадают.\nЭто может привести к потере данных.\nПродолжить?",
                QMessageBox.Yes | QMessageBox.No
            )
            if reply != QMessageBox.Yes:
                return
        
        # Предупреждение о перезаписи существующего файла
        if os.path.exists(output_file):
            reply = QMessageBox.question(
                self, "Подтверждение",
                f"Файл '{output_name}' уже существует.\nПерезаписать?",
                QMessageBox.Yes | QMessageBox.No
            )
            if reply != QMessageBox.Yes:
                return
        
        # Блокируем интерфейс
        self.set_ui_enabled(False)
        self.progress_bar.setVisible(True)
        self.progress_bar.setValue(0)
        self.status_label.setText("Начало обработки...")
        
        # Устанавливаем цвет статуса в зависимости от формата
        if is_xlsx:
            self.status_label.setStyleSheet("color: #3498db; font-size: 14px;")
        else:
            self.status_label.setStyleSheet("color: #4CAF50; font-size: 14px;")
        
        # Создаем и запускаем поток конвертации
        self.converter_thread = ConverterThread(input_file, output_file, is_xlsx)
        self.converter_thread.progress.connect(self.update_progress)
        self.converter_thread.message.connect(self.update_status)
        self.converter_thread.finished.connect(self.conversion_finished)
        self.converter_thread.start()
    
    def set_ui_enabled(self, enabled):
        """Включает или выключает элементы интерфейса"""
        self.input_file_edit.setEnabled(enabled)
        self.browse_file_btn.setEnabled(enabled)
        self.xlsx_radio.setEnabled(enabled)
        self.csv_radio.setEnabled(enabled)
        self.output_dir_edit.setEnabled(enabled)
        self.browse_dir_btn.setEnabled(enabled)
        self.output_name_edit.setEnabled(enabled)
        self.convert_btn.setEnabled(enabled)
    
    def update_progress(self, value):
        """Обновление прогресс бара"""
        self.progress_bar.setValue(value)
    
    def update_status(self, message):
        """Обновление статуса"""
        self.status_label.setText(message)
    
    def conversion_finished(self, success, message):
        """Завершение конвертации"""
        # Разблокируем интерфейс
        self.set_ui_enabled(True)
        self.progress_bar.setVisible(False)
        self.check_convert_button()
        
        if success:
            self.status_label.setText("✅ Преобразование завершено успешно!")
            self.status_label.setStyleSheet("color: #4CAF50; font-size: 14px;")
            
            # Показываем детальное сообщение об успехе
            msg_box = QMessageBox(self)
            msg_box.setWindowTitle("Успешно")
            msg_box.setText("Файл успешно преобразован")
            msg_box.setInformativeText(message)
            msg_box.setIcon(QMessageBox.Information)
            
            # Добавляем кнопки
            open_btn = msg_box.addButton("Открыть папку", QMessageBox.ActionRole)
            ok_btn = msg_box.addButton("OK", QMessageBox.AcceptRole)
            msg_box.setDefaultButton(ok_btn)
            
            msg_box.exec()
            
            # Обработка нажатия кнопки "Открыть папку"
            if msg_box.clickedButton() == open_btn:
                output_dir = self.output_dir_edit.text()
                if os.path.exists(output_dir):
                    # Открываем папку в проводнике
                    if sys.platform == "win32":
                        os.startfile(output_dir)
                    elif sys.platform == "darwin":
                        os.system(f'open "{output_dir}"')
                    else:
                        os.system(f'xdg-open "{output_dir}"')
        
        else:
            self.status_label.setText("❌ Ошибка при преобразовании")
            self.status_label.setStyleSheet("color: #f44336; font-size: 14px;")
            
            # Показываем сообщение об ошибке
            QMessageBox.critical(self, "Ошибка", message)

def main():
    # Проверяем наличие необходимых библиотек
    try:
        import pandas
        import openpyxl
        from PySide6 import QtWidgets
    except ImportError as e:
        print(f"Ошибка: Не удалось импортировать необходимые библиотеки")
        print(f"Установите их с помощью команд:")
        print(f"pip install pandas openpyxl PySide6")
        print(f"Ошибка импорта: {e}")
        return
    
    app = QApplication(sys.argv)
    app.setStyle("Fusion")  # Современный стиль
    
    # Запускаем приложение
    window = ConverterApp()
    window.show()
    
    sys.exit(app.exec())

if __name__ == "__main__":
    main()